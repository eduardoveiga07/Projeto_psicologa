import unittest
import io
import zipfile
import openpyxl
from datetime import date, datetime
from decimal import Decimal
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.models import Base, Paciente, AgendaSessao, Despesa, ContratoHistorico, Usuario, Perfil, Frequencia, TipoContrato, Auditoria
from app.services.exportacao import gerar_exportacao_zip


class ExportacaoTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Banco SQLite em memória para testes de exportação rápidos
        cls.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(cls.engine)
        cls.Session = sessionmaker(bind=cls.engine)

    def setUp(self):
        self.session = self.Session()
        # Limpa tabelas
        self.session.query(Paciente).delete()
        self.session.query(AgendaSessao).delete()
        self.session.query(Despesa).delete()
        self.session.query(ContratoHistorico).delete()
        self.session.query(Usuario).delete()
        self.session.query(Auditoria).delete()
        self.session.commit()

        # Insere dados de teste
        # 1. Paciente
        p = Paciente(
            nome="Eduardo Teste",
            telefone="5511999998888",
            data_nascimento=date(1995, 5, 20),
            tipo_contrato=TipoContrato.MENSAL,
            valor_sessao=Decimal("150.00"),
            frequencia=Frequencia.SEMANAL,
            dias_semana="Segunda-feira",
            horario_atendimento="Segunda-feira=10:00 - 11:00",
            ativo_desde=date(2026, 6, 1)
        )
        self.session.add(p)
        self.session.commit()

        # 2. Contrato
        c = ContratoHistorico(
            id_paciente=p.id_paciente,
            vigente_de=date(2026, 6, 1),
            frequencia=Frequencia.SEMANAL,
            valor_sessao=Decimal("150.00")
        )
        self.session.add(c)

        # 3. Sessão
        s = AgendaSessao(
            id_paciente=p.id_paciente,
            data_hora_inicio=datetime(2026, 6, 15, 10, 0),
            data_hora_fim=datetime(2026, 6, 15, 11, 0),
            valor_sessao=Decimal("150.00")
        )
        self.session.add(s)

        # 4. Despesa
        d = Despesa(
            descricao="Aluguel",
            valor=Decimal("1200.00"),
            data_vencimento=date(2026, 6, 10),
            mes_referencia="2026-06",
            paga=False
        )
        self.session.add(d)

        # 5. Usuário
        u = Usuario(
            username="dona_export",
            nome="Dona Maria",
            senha_hash="hash_secreta_123",
            perfil=Perfil.DONA,
            ativo=True
        )
        self.session.add(u)
        self.session.commit()

    def tearDown(self):
        self.session.close()

    def test_gerar_exportacao_zip_sucesso(self):
        # Executa a geração do ZIP
        zip_bytes = gerar_exportacao_zip(self.session)
        self.assertIsInstance(zip_bytes, bytes)
        self.assertGreater(len(zip_bytes), 0)

        # Carrega o ZIP para checar os arquivos contidos
        zip_file = zipfile.ZipFile(io.BytesIO(zip_bytes))
        namelist = zip_file.namelist()

        # Verifica todos os arquivos obrigatórios
        expected_files = [
            "pacientes.xlsx",
            "sessoes.xlsx",
            "despesas.xlsx",
            "contratos_historico.xlsx",
            "usuarios.xlsx",
            "auditoria.xlsx",
            "README.txt"
        ]
        for f in expected_files:
            self.assertIn(f, namelist)

        # Valida que usuarios.xlsx não contém a senha_hash
        usuarios_data = zip_file.read("usuarios.xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(usuarios_data))
        ws = wb.active

        # Lê os cabeçalhos (primeira linha)
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("senha_hash", headers)
        self.assertIn("username", headers)

        # Valida que pacientes.xlsx não contém o id_paciente
        pacientes_data = zip_file.read("pacientes.xlsx")
        wb_pac = openpyxl.load_workbook(io.BytesIO(pacientes_data))
        ws_pac = wb_pac.active
        pac_headers = [cell.value for cell in ws_pac[1]]
        self.assertNotIn("id_paciente", pac_headers)
        self.assertIn("nome", pac_headers)

        # Valida conteúdo do README
        readme_text = zip_file.read("README.txt").decode("utf-8")
        self.assertIn("Relatório de Exportação de Dados do Consultório", readme_text)
        self.assertIn("auditoria.xlsx: Trilha de auditoria", readme_text)
        self.assertIn("últimos 12 meses", readme_text)


if __name__ == "__main__":
    unittest.main()
